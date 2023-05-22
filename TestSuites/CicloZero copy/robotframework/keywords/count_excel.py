"""
Load and create a pandas dataframe from an excel file
"""
import re
import datetime
import pandas as pd
import openpyxl


def _load_excel_file(file_path: str) -> pd.DataFrame:
    """
    Load an excel file and return a pandas dataframe. The first row is used as the header
    """
    return pd.read_excel(file_path, header=0)

def _load_tsv_file(file_path: str) -> pd.DataFrame:
    """
    Load a tsv file and return a pandas dataframe. The first row is used as the header
    """
    return pd.read_csv(file_path, sep="\t", header=0)

def _count_number_of_same_values_in_column(dataframe: pd.DataFrame, column_name: str) -> pd.Series:
    """
    Count the number of same values in a column
    """
    return dataframe[column_name].value_counts()

def _add_empty_column_to_dataframe(dataframe: pd.DataFrame, column_name: str):
    """
    Add an empty column to a pandas dataframe
    """
    dataframe[column_name] = ""

def _save_dataframe_to_excel(dataframe: pd.DataFrame, file_path: str):
    """
    Save a pandas dataframe to an excel file
    """
    dataframe.to_excel(file_path, index=False)

def _add_total_column_to_dataframe(dataframe: pd.DataFrame, column_name: str):
    """
    Add a column to a pandas dataframe with the formula "=B2-C2-D2-E2"
    """
    for i in range(len(dataframe)):
        dataframe.loc[i, column_name] = f"=B{i+2}-C{i+2}-D{i+2}-E{i+2}"

def _add_attributes_columns(dataframe: pd.DataFrame):
    """
    Add columns to a pandas dataframe with the attributes from the "prod" column.
    Add the columns if not exists
    """
    for column in ["id_modelo", "color", "almacenamiento", "calidad", "estado", "reacondicionado"]:
        if column not in dataframe.columns:
            _add_empty_column_to_dataframe(dataframe, column)

    for i in range(len(dataframe)):
        prod = str(dataframe.loc[i, "prod"])
        if "[" in prod:
            sku = prod[prod.find("[") + 1:prod.find("]")]
        else:
            sku = prod
        attributes = get_attributes_from_sku(sku)
        dataframe.loc[i, "id_modelo"] = attributes["id_modelo"]
        dataframe.loc[i, "color"] = attributes["color"]
        dataframe.loc[i, "almacenamiento"] = attributes["almacenamiento"]
        dataframe.loc[i, "calidad"] = attributes["calidad"]
        dataframe.loc[i, "estado"] = attributes["estado"]
        dataframe.loc[i, "reacondicionado"] = attributes["reacondicionado"]

def get_attributes_from_sku(sku: str) -> dict[str, str]:
    """
    From SKU: [<id_modelo>-<color(SL=Silver)>-<almacenamiento(128)>-<calidad(A,B,B+,C)> -<reacondicionado>]
    Get the attributes: {"id_modelo": <id_modelo>, "color": <color>, "almacenamiento": <almacenamiento>, "calidad": <calidad>, "reacondicionado": <reacondicionado>}
    """
    gropus_estado = {
        "A+": "NUEVO",
        "A": "NUEVO",
        "B+": "NUEVO",
        "B": "NUEVO",
        "C": "MEDIO",
        "D": "BAJO",
        "ASIS": "IGNORE"
    }

    attributes = sku.split("-")
    assert attributes[3].strip().upper() in gropus_estado.keys(), f"Estado {attributes[3].strip().upper()} not in {gropus_estado.keys()}"

    attributes_dict = {
        "id_modelo": attributes[0].upper(),
        "color": attributes[1].upper(),
        "almacenamiento": attributes[2].upper(),
        "calidad": attributes[3].strip().upper(),
        "estado": gropus_estado[attributes[3].strip().upper()],
        "reacondicionado": attributes[4].upper() if len(attributes) == 5 else "",
    }
    return attributes_dict

# KEYORDS

def create_excel(count_excel_path: str, output_excel_path: str):
    """
    Create an excel file with the count of the same values in a column:
    Columns: "prod", "count", "amz unshipped", "amz pending", "flend", "Total", "id_modelo", "color", "almacenamiento", "calidad", "reacondicionado"
    """

    df = _load_excel_file(count_excel_path)
    model_count = _count_number_of_same_values_in_column(df, "Producto")

    # Convert Series[int] to DataFrame[int] with columns "Producto" and "Cantidad"
    model_count = model_count.to_frame().reset_index()
    model_count.columns = ["prod", "count"]

    # Add empty column "amz unshipped", "amz pending" "flend"
    _add_empty_column_to_dataframe(model_count, "amz unshipped")
    _add_empty_column_to_dataframe(model_count, "amz pending")
    _add_empty_column_to_dataframe(model_count, "flend")

    # Add column "Total" with "Cantidad" - "amz unshipped" - "amz pending" - "flend". Values must be calculated with excel formulas (=Bi-Ci-Di-Ei)
    _add_total_column_to_dataframe(model_count, "Total")
    _add_attributes_columns(model_count)
    _save_dataframe_to_excel(model_count, output_excel_path)

    return model_count


def append_tsv_to_main_excel(tsv_path: str, excel_path:str, output_excel_path: str):
    excel_df = _load_excel_file(excel_path)
    tsv_df = _load_tsv_file(tsv_path)

    model_count = _count_number_of_same_values_in_column(tsv_df, "sku")

    model_count = model_count.to_frame().reset_index()
    model_count.columns = ["prod", "count"]

    for i in range(len(model_count)):
        prod = model_count.loc[i, "prod"]
        pattern = re.escape(str(prod))
        excel_row = excel_df[excel_df["prod"].str.contains(pattern)]
        if excel_row.empty:
            # Match id_modelo, color, almacenamiento, estado, reacondicionado
            attributes = get_attributes_from_sku(str(prod))
            matched_id_modelo = excel_df[excel_df["id_modelo"] == attributes["id_modelo"]]
            matched_id_color = matched_id_modelo[matched_id_modelo["color"] == attributes["color"]]
            matched_id_almacenamiento = matched_id_color[matched_id_color["almacenamiento"] == attributes["almacenamiento"]]
            matched_id_calidad = matched_id_almacenamiento[matched_id_almacenamiento["estado"] == attributes["estado"]]
            matched_id_reacondicionado = matched_id_calidad[matched_id_calidad["reacondicionado"] == attributes["reacondicionado"]]

            if len(matched_id_reacondicionado) == 1:
                print(f'Sku "{prod}" matched with "{matched_id_reacondicionado.iloc[0]["prod"]}" ignoring calidad')
                excel_df.loc[matched_id_reacondicionado.index, "amz unshipped"] = model_count.loc[i, "count"]  # type: ignore
            elif len(matched_id_reacondicionado) > 1:
                # Assign to the one with calidad = A
                print(f'Sku "{prod}" matched with "{matched_id_reacondicionado.iloc[0]["prod"]}" ignoring calidad')
                excel_df.loc[ matched_id_calidad[matched_id_calidad["calidad"] == 'A'], "amz unshipped"] = model_count.loc[i, "count"]  # type: ignore
            else:
            # No match
                new_row = [prod, "", model_count.loc[i, "count"]] + [""] * (len(excel_df.columns) - 3)
                excel_df.loc[len(excel_df)] = new_row # type: ignore
        else:
            excel_df.loc[excel_row.index, "amz unshipped"] = model_count.loc[i, "count"] # type: ignore

    _add_total_column_to_dataframe(excel_df, "Total")
    _add_attributes_columns(excel_df)
    _save_dataframe_to_excel(excel_df, output_excel_path)

    return excel_df


def append_dict_to_main_excel(order_count_dict: dict, excel_path:str, output_excel_path: str):
    excel_df = _load_excel_file(excel_path)

    for prod, count in order_count_dict.items():
        pattern = re.escape(str(prod))
        excel_row = excel_df[excel_df["prod"].str.contains(pattern)]
        if excel_row.empty:
             # Match id_modelo, color, almacenamiento, estado, reacondicionado
            attributes = get_attributes_from_sku(str(prod))
            matched_id_modelo = excel_df[excel_df["id_modelo"] == attributes["id_modelo"]]
            matched_id_color = matched_id_modelo[matched_id_modelo["color"] == attributes["color"]]
            matched_id_almacenamiento = matched_id_color[matched_id_color["almacenamiento"] == int(attributes["almacenamiento"])]
            matched_id_calidad = matched_id_almacenamiento[matched_id_almacenamiento["estado"] == attributes["estado"]]
            matched_id_reacondicionado = matched_id_calidad[matched_id_calidad["reacondicionado"] == attributes["reacondicionado"]]

            if len(matched_id_reacondicionado) == 1:
                print(f'Sku "{prod}" matched with "{matched_id_reacondicionado.iloc[0]["prod"]}" ignoring calidad')
                excel_df.loc[matched_id_reacondicionado.index, "amz pending"] = count
            elif len(matched_id_reacondicionado) > 1:
                # Assign to the one with calidad = A
                print(f'Sku "{prod}" matched with "{matched_id_reacondicionado.iloc[0]["prod"]}" ignoring calidad')
                excel_df.loc[matched_id_calidad[matched_id_calidad["calidad"] == 'A'].index, "amz pending"] = count
            else:
                new_row = [prod, "", "", count] + [""] * (len(excel_df.columns) - 4)
                excel_df.loc[len(excel_df)] = new_row # type: ignore
        else:
            excel_df.loc[excel_row.index, "amz pending"] = count # type: ignore

    _add_total_column_to_dataframe(excel_df, "Total")
    _add_attributes_columns(excel_df)
    _save_dataframe_to_excel(excel_df, output_excel_path)

    return excel_df


def create_sheet_for_sku(sku: str, output_excel_path: str):
    wb = openpyxl.Workbook()

    # Check if sheet exists
    if sku in wb.sheetnames:
        print(f"Sheet {sku} already exists")
        return
    else:
        # Create sheet
        wb.create_sheet(sku)
    
    # Add columns: marketplace, status, self price, best price, best seller, second price, second seller, third price, third seller, url, timestamp
    sheet = wb[sku]
    sheet["A1"] = "marketplace"
    sheet["B1"] = "status"
    sheet["C1"] = "self price"
    sheet["D1"] = "best price"
    sheet["E1"] = "best seller"
    sheet["F1"] = "second price"
    sheet["G1"] = "second seller"
    sheet["H1"] = "third price"
    sheet["I1"] = "third seller"
    sheet["J1"] = "url"
    sheet["K1"] = "timestamp"

    # Save
    wb.save(output_excel_path)


def add_prices_by_sku_and_market(excel_path: str, sku: str, marketplace: str, status: str, self_price: str, prices: dict, url: str):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[sku]

    # Check if sheet exists
    if sku not in wb.sheetnames:
        print(f"Sheet {sku} does not exists")
        return

    # Add row with: marketplace, status, self price, best price, best seller, second price, second seller, third price, third seller, url
    row = sheet.max_row + 1
    sheet[f"A{row}"] = marketplace
    sheet[f"B{row}"] = status
    sheet[f"C{row}"] = self_price
    # Best 3 prices. Sort dict by value
    sorted_prices = {k: v for k, v in sorted(prices.items(), key=lambda item: item[1])}

    for i, (seller, price) in enumerate(sorted_prices.items()):
        sheet[f"{chr(ord('D') + (i * 2))}{row}"] = price
        sheet[f"{chr(ord('E') + (i * 2))}{row}"] = seller
        if i == 2:
            break
    sheet[f"J{row}"] = url
    sheet[f"K{row}"] = datetime.datetime.now().strftime("%H:%M %d/%m/%Y")

    # Save
    wb.save(excel_path)

def add_label_by_sku(excel_path: str, sku: str, marketplace: str, status: str, label: str, self_price:str, url: str):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[sku]

    # Check if sheet exists
    if sku not in wb.sheetnames:
        print(f"Sheet {sku} does not exists")
        return

    # Add row with: marketplace, status, self price, best price, best seller, second price, second seller, third price, third seller, url
    row = sheet.max_row + 1
    sheet[f"A{row}"] = marketplace
    sheet[f"B{row}"] = status
    sheet[f"C{row}"] = self_price
    # Add 'not found' to all prices
    for i in range(3):
        sheet[f"{chr(ord('D') + (i * 2))}{row}"] = label
        sheet[f"{chr(ord('E') + (i * 2))}{row}"] = label
    sheet[f"J{row}"] = url

    # Save
    wb.save(excel_path)


if __name__ == "__main__":
    import os 

    workdir = "/workspaces/ai-butlerhat/data-butlerhat/robotframework-butlerhat/TestSuites/CicloZero/downloads"
    stock_excel_path = os.path.join(workdir, "stock.quant.xlsx")
    res_stock_excel_path = os.path.join(workdir, "deb.stock.quant.result.xlsx")
    tsv_path = os.path.join(workdir, "unshipped.tsv")
    pending_unshipped_path = os.path.join(workdir, "stock.quant.amz.unshipped.result.xlsx")
    res_full_excel_path = os.path.join(workdir, "deb.full.result.xlsx")

    # create_excel(stock_excel_path, res_stock_excel_path)
    # append_tsv_to_main_excel(tsv_path, res_stock_excel_path, res_full_excel_path)
    df = append_dict_to_main_excel({
        'iPXS-GD-64-B+ -R': 1, 
        'iP12MN-WT-128-B -R': 1, 
        'iPXS-SL-256-A -R': 1, 
        'iP12PR-SL-128-A -R': 3, 
        'iP13-PK-128-A -R': 1, 
        'iP13-BL-128-B -R': 1, 
        'iP12PR-SL-128-B -R': 1, 
        'iP12PR-GRT-128-A -R': 1
        }, pending_unshipped_path, res_full_excel_path)

    print(f"Excel file saved in {res_full_excel_path}")